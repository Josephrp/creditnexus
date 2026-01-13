"""
Audit Export Service for CreditNexus.

Provides export functionality for audit data in various formats (CSV, Excel, PDF).
"""

import logging
import csv
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
from sqlalchemy.orm import Session

from app.services.audit_service import AuditService
from app.db.models import AuditLog

logger = logging.getLogger(__name__)


class AuditExportService:
    """Service for exporting audit data to various formats."""
    
    def __init__(self):
        """Initialize audit export service."""
        self.audit_service = AuditService()
    
    def export_to_csv(
        self,
        db: Session,
        audit_logs: List[AuditLog],
        include_metadata: bool = True
    ) -> bytes:
        """
        Export audit logs to CSV format.
        
        Args:
            db: Database session
            audit_logs: List of AuditLog instances
            include_metadata: Whether to include metadata column
            
        Returns:
            CSV file as bytes
        """
        try:
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            headers = [
                "ID", "User ID", "User Name", "User Email", "Action",
                "Target Type", "Target ID", "IP Address", "User Agent",
                "Occurred At"
            ]
            if include_metadata:
                headers.append("Metadata")
            
            writer.writerow(headers)
            
            # Write rows
            for log in audit_logs:
                enriched = self.audit_service.enrich_audit_log(db, log)
                row = [
                    log.id,
                    log.user_id,
                    enriched.get("user", {}).get("name") if enriched.get("user") else None,
                    enriched.get("user", {}).get("email") if enriched.get("user") else None,
                    log.action,
                    log.target_type,
                    log.target_id,
                    log.ip_address,
                    log.user_agent,
                    log.occurred_at.isoformat() if log.occurred_at else None,
                ]
                if include_metadata:
                    import json
                    row.append(json.dumps(log.action_metadata) if log.action_metadata else "")
                
                writer.writerow(row)
            
            # Convert to bytes
            output.seek(0)
            return output.getvalue().encode('utf-8')
            
        except Exception as e:
            logger.error(f"Failed to export to CSV: {e}", exc_info=True)
            raise
    
    def export_to_excel(
        self,
        db: Session,
        audit_logs: List[AuditLog],
        include_metadata: bool = True
    ) -> bytes:
        """
        Export audit logs to Excel format with multiple sheets.
        
        Args:
            db: Database session
            audit_logs: List of AuditLog instances
            include_metadata: Whether to include metadata column
            
        Returns:
            Excel file as bytes
        """
        try:
            try:
                import pandas as pd
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
            except ImportError:
                logger.error("pandas and openpyxl required for Excel export")
                raise ValueError("Excel export requires pandas and openpyxl packages")
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Sheet 1: Audit Logs
            ws_logs = wb.create_sheet("Audit Logs")
            
            # Prepare data
            data = []
            for log in audit_logs:
                enriched = self.audit_service.enrich_audit_log(db, log)
                row = {
                    "ID": log.id,
                    "User ID": log.user_id,
                    "User Name": enriched.get("user", {}).get("name") if enriched.get("user") else None,
                    "User Email": enriched.get("user", {}).get("email") if enriched.get("user") else None,
                    "Action": log.action,
                    "Target Type": log.target_type,
                    "Target ID": log.target_id,
                    "IP Address": log.ip_address,
                    "User Agent": log.user_agent,
                    "Occurred At": log.occurred_at.isoformat() if log.occurred_at else None,
                }
                if include_metadata:
                    import json
                    row["Metadata"] = json.dumps(log.action_metadata) if log.action_metadata else ""
                
                data.append(row)
            
            # Write to sheet
            df = pd.DataFrame(data)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_logs.append(r)
            
            # Sheet 2: Summary Statistics
            ws_summary = wb.create_sheet("Summary")
            
            # Count by action
            action_counts = {}
            for log in audit_logs:
                action_counts[log.action] = action_counts.get(log.action, 0) + 1
            
            summary_data = [
                ["Metric", "Value"],
                ["Total Logs", len(audit_logs)],
                ["Unique Users", len(set(log.user_id for log in audit_logs if log.user_id))],
                ["Date Range", f"{min(log.occurred_at for log in audit_logs if log.occurred_at).isoformat() if audit_logs else 'N/A'} to {max(log.occurred_at for log in audit_logs if log.occurred_at).isoformat() if audit_logs else 'N/A'}"],
                ["", ""],
                ["Action", "Count"],
            ]
            
            for action, count in sorted(action_counts.items(), key=lambda x: x[1], reverse=True):
                summary_data.append([action, count])
            
            for row in summary_data:
                ws_summary.append(row)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Failed to export to Excel: {e}", exc_info=True)
            raise
    
    def export_to_pdf(
        self,
        db: Session,
        audit_logs: List[AuditLog],
        title: str = "Audit Report"
    ) -> bytes:
        """
        Export audit logs to PDF format.
        
        Args:
            db: Database session
            audit_logs: List of AuditLog instances
            title: Report title
            
        Returns:
            PDF file as bytes
        """
        try:
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib.units import inch
            except ImportError:
                logger.error("reportlab required for PDF export")
                raise ValueError("PDF export requires reportlab package")
            
            # Create PDF document
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            story.append(Paragraph(title, styles['Title']))
            story.append(Spacer(1, 0.2 * inch))
            
            # Summary
            story.append(Paragraph("Summary", styles['Heading2']))
            summary_data = [
                ["Total Logs", str(len(audit_logs))],
                ["Unique Users", str(len(set(log.user_id for log in audit_logs if log.user_id)))],
            ]
            
            if audit_logs:
                dates = [log.occurred_at for log in audit_logs if log.occurred_at]
                if dates:
                    summary_data.append([
                        "Date Range",
                        f"{min(dates).isoformat()} to {max(dates).isoformat()}"
                    ])
            
            summary_table = Table(summary_data, colWidths=[2 * inch, 4 * inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.3 * inch))
            
            # Audit Logs Table (limited to first 100 for PDF)
            story.append(Paragraph("Audit Logs (showing first 100)", styles['Heading2']))
            
            # Prepare table data
            table_data = [["ID", "User", "Action", "Target", "Occurred At"]]
            
            for log in audit_logs[:100]:  # Limit to 100 rows for PDF
                enriched = self.audit_service.enrich_audit_log(db, log)
                user_name = enriched.get("user", {}).get("name") if enriched.get("user") else "Unknown"
                target = f"{log.target_type}#{log.target_id}" if log.target_type and log.target_id else "N/A"
                occurred_at = log.occurred_at.isoformat() if log.occurred_at else "N/A"
                
                table_data.append([
                    str(log.id),
                    user_name,
                    log.action,
                    target,
                    occurred_at
                ])
            
            # Create table
            table = Table(table_data, colWidths=[0.5 * inch, 1.5 * inch, 1.5 * inch, 2 * inch, 2.5 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            story.append(table)
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Failed to export to PDF: {e}", exc_info=True)
            raise
    
    def export_audit_log_entry(
        self,
        db: Session,
        audit_log: AuditLog,
        format: str = "json"
    ) -> bytes:
        """
        Export a single audit log entry.
        
        Args:
            db: Database session
            audit_log: AuditLog instance
            format: Export format ('json', 'csv')
            
        Returns:
            Exported data as bytes
        """
        try:
            enriched = self.audit_service.enrich_audit_log(db, audit_log)
            
            if format == "json":
                import json
                return json.dumps(enriched, indent=2, default=str).encode('utf-8')
            
            elif format == "csv":
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(["Field", "Value"])
                for key, value in enriched.items():
                    if isinstance(value, dict):
                        for sub_key, sub_value in value.items():
                            writer.writerow([f"{key}.{sub_key}", sub_value])
                    else:
                        writer.writerow([key, value])
                output.seek(0)
                return output.getvalue().encode('utf-8')
            
            else:
                raise ValueError(f"Unsupported format: {format}")
                
        except Exception as e:
            logger.error(f"Failed to export audit log entry: {e}", exc_info=True)
            raise
